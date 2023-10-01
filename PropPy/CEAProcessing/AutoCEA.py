# AutoCEA
# By Hud & GPT

import pandas as pd
import numpy as np
import sys
import math
from CEATabulationCleaner import CEATabCleaner
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options

print("\n\nStarting 'AutoCEA.py'.\n")
print("Reading CEA inputs from AutoCEA_Input.xlsx.\n")

######################### INPUTS ########################

# Read the Excel file / Handle Errors
try:
    pd.read_excel('CEAProcessing\AutoCEA_Input.xlsx')
except (PermissionError):
    print("Error: Close Input file and run again.\n")
    print("END\n\n")
    sys.exit()
ReadExcel = pd.ExcelFile('CEAProcessing\AutoCEA_Input.xlsx')

# Get the list of sheet names
the_sheet_names = ReadExcel.sheet_names

# Determine which sheet to pull parameters from
SheetBox = False
SheetCounter = 0 # variable to switch thru sheets

while SheetBox == False:

    sheet_name = SheetCounter # start at first sheet

    # Read the Excel file
    ReadExcel = pd.read_excel('CEAProcessing\AutoCEA_Input.xlsx', sheet_name)

    # Read check box value
    SheetBox = ReadExcel.iloc[1, 6] 

    if SheetBox == True:
        RunSheet = SheetCounter
        SheetDesc = ReadExcel.iloc[0, 3]
    else:
        SheetCounter += 1
print(f'Sheet to use: {RunSheet}, Sheet {SheetDesc}' )

# store CEA Data Identification Key
ReadExcel = pd.read_excel('CEAProcessing\AutoCEA_Input.xlsx', RunSheet)
CEA_Key = ReadExcel.iloc[2, 3] 
print(f'Data Identification Key: {CEA_Key}')

# store Chamber Pressure
Pc = ReadExcel.iloc[6, 3] 
print(f'Chamber Pressure: {Pc}')

# store Fuel
FuelBox = False
FuelCounter = 10

while FuelBox == False:

    # Read check box value
    FuelBox = ReadExcel.iloc[FuelCounter, 4] 

    if FuelBox == True:
        RunFuel = ReadExcel.iloc[FuelCounter, 3]
        FuelTemp = ReadExcel.iloc[17, 3]
    else:
        FuelCounter += 1
print(f'Fuel to use: {RunFuel} at {FuelTemp} K')

# store Oxidizer
OxBox = False
OxCounter = 21

while OxBox == False:

    # Read check box value
    OxBox = ReadExcel.iloc[OxCounter, 4] 

    if OxBox == True:
        RunOx = ReadExcel.iloc[OxCounter, 3]
        OxTemp = ReadExcel.iloc[27, 3]
    else:
        OxCounter += 1
print(f'Oxidizer to use: {RunOx} at {OxTemp} K')

# store MR
MRCounter = 6
MRArray = [None]*30

# Read check box value
MRBox = ReadExcel.iloc[31, 3] 

if MRBox == True:
    while MRCounter < 36:
        MRArray[MRCounter-6] = float(ReadExcel.iloc[34, MRCounter]) 
        if np.isnan(MRArray[MRCounter-6]):
            MRArray[MRCounter-6] = 0.0
        MRCounter += 1
    print("List of MRs:")
    print(MRArray)
else:
    MRLow = ReadExcel.iloc[33, 3]
    MRHigh = ReadExcel.iloc[34, 3]
    MRInt = ReadExcel.iloc[36, 3]
    print(f'MR Low: {MRLow}, MR High: {MRHigh}, MR Interval: {MRInt},')

# Store Exit Conditions
Pc_Pe = round(ReadExcel.iloc[41, 3],3) 
print(f'Pc/Pe: {Pc_Pe}')

print("\nInputs stored, beginning CEA macro.\n")
########################## CEARUN ##############################

#Open Firefox, go to CEA
driver = webdriver.Firefox()
driver.get("https://cearun.grc.nasa.gov/")

#Input Data Identification Key and submit
DIKField = driver.find_element(By.NAME, "tmpID")
DIKField.send_keys(CEA_Key)
DIKSubmit = driver.find_element(By.NAME, "Submit")
DIKSubmit.click()

#Complete Pressures Tab and Submit
PresField = driver.find_element(By.NAME, "P1")
PresField.send_keys(Pc)
PresRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[1]/input[4]")
PresRadio.click()
PresSubmit = driver.find_element(By.NAME, ".submit")
PresSubmit.click()

#Complete Fuels and Submit
if RunFuel == "CH4 (L)":
    FuelRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[2]")
    FuelRadio.click()
elif RunFuel == "H2":
    FuelRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[3]")
    FuelRadio.click()
elif RunFuel == "H2 (L)":
    FuelRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[4]")
    FuelRadio.click()
elif RunFuel == "RP-1":
    FuelRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[5]")
    FuelRadio.click()
elif RunFuel == "C2H6O":
    FuelRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[6]")
    FuelRadio.click()

FuelField = driver.find_element(By.NAME, "TsimpleFuel") #Enter Temp
FuelField.send_keys(FuelTemp)
FuelSubmit1 = driver.find_element(By.NAME, ".submit")
FuelSubmit1.click()

if RunFuel == "C2H6O": #Periodic Table Selection
    FuelC = driver.find_element(By.XPATH, '//*[@id="C"]')
    FuelC.click()
    FuelH = driver.find_element(By.XPATH, '//*[@id="H"]')
    FuelH.click()
    FuelO = driver.find_element(By.XPATH, '//*[@id="O"]')
    FuelO.click()
    FuelSubmit2 = driver.find_element(By.NAME, ".submit")
    FuelSubmit2.click()
    EthBox = driver.find_element(By.XPATH, "/html/body/form/fieldset/table/tbody/tr[1]/td[2]/input")
    EthBox.click()
    FuelSubmit3 = driver.find_element(By.NAME, ".submit")
    FuelSubmit3.click()    
    FuelSubmit4 = driver.find_element(By.NAME, ".submit")
    FuelSubmit4.click()  
    FuelSubmit5 = driver.find_element(By.NAME, ".submit")
    FuelSubmit5.click()
    FuelSubmit6 = driver.find_element(By.NAME, ".submit")
    FuelSubmit6.click()

#Complete Ox and Submit
if RunOx == "Air":
    OxRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[1]")
    OxRadio.click()
elif RunOx == "H2O2 (L)":
    OxRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[6]")
    OxRadio.click()
elif RunOx == "N2O":
    OxRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[8]")
    OxRadio.click()
elif RunOx == "O2":
    OxRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[10]")
    OxRadio.click()
elif RunOx == "O2 (L)":
    OxRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset/p[2]/input[11]")
    OxRadio.click()

OxField = driver.find_element(By.NAME, "TsimpleOxid") #Enter Temp
OxField.send_keys(OxTemp)
OxSubmit1 = driver.find_element(By.NAME, ".submit")
OxSubmit1.click()

#MRs
MRBoxCounter = 0
if MRBox == True:
    while MRBoxCounter < len(MRArray):
        MRPathName = "OFP" + str(MRBoxCounter+1)
        MRIndivField = driver.find_element(By.NAME, f'{MRPathName}')
        MRIndivField.send_keys(MRArray[MRBoxCounter])
        MRBoxCounter += 1
else:
    MRLowField = driver.find_element(By.NAME, "OFP_low")
    MRLowField.send_keys(MRLow)
    MRHighField = driver.find_element(By.NAME, "OFP_hi")
    MRHighField.send_keys(MRHigh)
    MRIntField = driver.find_element(By.NAME, "OFP_int")
    MRIntField.send_keys(MRInt)

MRSubmit = driver.find_element(By.NAME, ".submit")
MRSubmit.click()

#Exit Cond
ExitCondField = driver.find_element(By.NAME, "pip1") #Enter Temp
ExitCondField.send_keys(Pc_Pe)
ExCoSubmit = driver.find_element(By.NAME, ".submit")
ExCoSubmit.click()

#Final
IonsRadio = driver.find_element(By.NAME, "ions")
if not(IonsRadio.is_selected()):
    IonsRadio.click()
TabulateRadio = driver.find_element(By.XPATH, "/html/body/form/fieldset[5]/table/tbody/tr[1]/td[2]/input")
TabulateRadio.click()

FinalSubmit = driver.find_element(By.NAME, ".submit")
FinalSubmit.click()

#Spreadsheet Export Setup
PropsField1 = driver.find_element(By.NAME, "plt1")
PropsField1.send_keys("rho")
PropsField2 = driver.find_element(By.NAME, "plt2")
PropsField2.send_keys("gam")
PropsField3 = driver.find_element(By.NAME, "plt3")
PropsField3.send_keys("m")
PropsField4 = driver.find_element(By.NAME, "plt4")
PropsField4.send_keys("p")
PropsField5 = driver.find_element(By.NAME, "plt5")
PropsField5.send_keys("t")
PropsField6 = driver.find_element(By.NAME, "plt6")
PropsField6.send_keys("isp")
PropsField7 = driver.find_element(By.NAME, "plt7")
PropsField7.send_keys("mach")
PropsField8 = driver.find_element(By.NAME, "plt8")
PropsField8.send_keys("son")
PropsSubmit = driver.find_element(By.NAME, ".submit")
PropsSubmit.click()

#Collect Output
OutputLink = driver.find_element(By.LINK_TEXT, 'Output')
OutputLink.click()

#Send Output source to text file
CopyOutputText = driver.page_source
RawCEA_Output_File = 'CEAProcessing\RawCEA_Output.txt'
with open(RawCEA_Output_File, 'w', encoding='utf-8') as file:
    file.write(CopyOutputText)

#Go back
driver.back()

#Collect Tabulation
TabulationLink = driver.find_element(By.LINK_TEXT, 'Tabulation')
TabulationLink.click()

print("CEA macro complete, sending CEA Output and Tabulation to raw text files.\n")
print("Sending raw Tabulation to 'CEATabulationCleaner.py'.\n")

#Send Tabulation to text file
CopyTabText = driver.page_source
RawCEA_Tab_File = 'CEAProcessing\RawCEA_Tabulation.txt'
with open(RawCEA_Tab_File, 'w', encoding='utf-8') as file:
    file.write(CopyTabText)

#CEA OutPut Excel
try: #handle open error
    openpyxl.load_workbook('CEAProcessing\AutoCEA_Output.xlsx')
except (PermissionError):
    print("\nError: Close Output file and run again.\n")
    print("END\n\n")
    sys.exit()

OutputExcelFile = openpyxl.load_workbook('CEAProcessing\AutoCEA_Output.xlsx')
ExistingSheetNames = OutputExcelFile.sheetnames
NewSheetNameAttempt = CEA_Key
SheetNameCounter = 1

while NewSheetNameAttempt in ExistingSheetNames:
    NewSheetNameAttempt = CEA_Key
    NewSheetNameAttempt = NewSheetNameAttempt + f"({SheetNameCounter})"
    SheetNameCounter += 1

OutputExcelFile.create_sheet(NewSheetNameAttempt)

# Start Sheet Edit (NewSheetEdit is the active sheet name)

#TitleBlock
NewSheetEdit = OutputExcelFile[NewSheetNameAttempt]
NewSheetEdit.column_dimensions['A'].width = 20/7
NewSheetEdit.column_dimensions['B'].width = 17.14
NewSheetEdit['B2'] = "AutoCEA Output"
NewSheetEdit['B2'].alignment = Alignment(horizontal='center', vertical='center')
NewSheetEdit['B2'].font = Font(bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
NewSheetEdit['B2'].border = border
blue_fill = PatternFill(start_color="DDE1EC", end_color="DDE1EC", fill_type="solid")
NewSheetEdit['B2'].fill = blue_fill
NewSheetEdit.sheet_view.zoomScale = 85

# Make Tabulation Excel ready
CEATabCleaner("CEAProcessing\RawCEA_Tabulation.txt","CEAProcessing\CleanCEA_Tabulation.txt")
print("Cleaned Tabulation saved to 'CleanCEA_Tabulation.txt'.\n")

print("Sending Output and clean Tabulation to 'AutoCEA_Output.xlsx'.\n")

#Paste CEA Output and Tabulation From Text Files into Excel
with open(RawCEA_Output_File, 'r') as file:
    text_lines = file.readlines()[7:]
for index, line in enumerate(text_lines, start=4):
    NewSheetEdit.cell(row=index, column=2, value=line)

# Read CSV data
CleanCEA_Tab_File = "CEAProcessing\CleanCEA_Tabulation.txt"
df2 = pd.read_csv(CleanCEA_Tab_File, header=None)

# Choose the cell where you want to paste the data
Tab_start_row = 3  # Row where you want to start pasting data
Tab_start_column = 13  # Column where you want to start pasting data

# Write the DataFrame values to the selected cells
for r_idx, row in enumerate(df2.iterrows(), start=Tab_start_row):
    for c_idx, value in enumerate(row[1], start=Tab_start_column):
        cell = NewSheetEdit.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.number_format = '0.0000E+00'
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Add back titles to data
TabTitles = ["rho", "gam", "m", "p", "t", "isp", "mach", "son"]
Title_start_column = "M"
row = 2
for idx, value in enumerate(TabTitles):
    cell = f"{Title_start_column}{row}"
    NewSheetEdit[cell].value = value
    NewSheetEdit[cell].alignment = Alignment(horizontal='center', vertical='center')
    NewSheetEdit.column_dimensions[Title_start_column].width = 11
    Title_start_column = chr(ord(Title_start_column) + 1)

# Define the columns to apply left and right borders (L through Q)
start_column = "L"
end_column = "S"

# Define the start and end rows
with open(CleanCEA_Tab_File, 'r') as file:
    line_count = sum(1 for line in file)
start_row = 2  # Change to the desired start row
end_row = line_count + 1   # Change to the desired end row

# Iterate through the specified rows and columns and add left and right borders
for row_num in range(start_row, end_row + 2):
    for column in NewSheetEdit.iter_cols(min_col=openpyxl.utils.column_index_from_string(start_column),
                                  max_col=openpyxl.utils.column_index_from_string(end_column),
                                  min_row=row_num, max_row=row_num):
        for cell in column:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin')
            )

# Define the border style you want to apply
border_style = openpyxl.styles.Side(border_style='thin')

# Add left, bottom, and right borders to the cell
NewSheetEdit['K2'].border = openpyxl.styles.Border(
    right=border_style,
    bottom=border_style
)
NewSheetEdit['L2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['M2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['N2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['O2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['P2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['Q2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['R2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['S2'].border = openpyxl.styles.Border(
    left=border_style,
    right=border_style,
    bottom=border_style
)
NewSheetEdit['T2'].border = openpyxl.styles.Border(
    left=border_style,
    bottom=border_style
)

#Adding words and colors

# Define the words to repeat
words = ["Chamber", "Throat", "Exit"]

# Define the colors for each row
row_colors = [
    "FCD9D1",  # Row 3 color
    "FCF7D1",  # Row 4 color
    "E0D1FC",  # Row 5 color
]

# Define the starting row and column
start_row1 = 3
start_row2 = 3
start_column = 11  # Column K

# Repeat the words 10 times
for i in range(int(line_count/3)):
    for word in words:
        cell = NewSheetEdit.cell(row=start_row1, column=start_column)
        cell.value = word
        start_row1 += 1  # Move to the next row
    for row_color in row_colors:
        # Iterate through columns K to T (11 to 20)
        for column_index in range(start_column, start_column + 10):
            cell = NewSheetEdit.cell(row=start_row2, column=column_index)
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

        start_row2 += 1  # Move to the next row

NewSheetEdit['K2'] = "Location"
NewSheetEdit['K2'].alignment = Alignment(horizontal='center', vertical='center')
NewSheetEdit['L2'] = "MR"
NewSheetEdit.column_dimensions["L"].width = 6

# Specify the column you want to left-align (e.g., column 'A')
column_letter = 'L'

# Get the maximum row number in the column
max_row = NewSheetEdit.max_row

# Iterate through the rows in the specified column and set the alignment
for row_number in range(1, max_row + 1):
    cell = NewSheetEdit[f'{column_letter}{row_number}']
    cell.alignment = Alignment(horizontal='left')

NewSheetEdit['L2'].alignment = Alignment(horizontal='center', vertical='center')

# Create a list for the MRs depending on input selection
if MRBox == True:
    MRIndex = 0
    MRIntArray = []
    while MRIndex < len(MRArray):
        if MRArray[MRIndex] > 0:
            MRIntArray.append(MRArray[MRIndex])
        MRIndex += 1
    values_list = MRIntArray  # Replace with your list of values

    # Define the starting row and column
    start_row = 3  # Row K3
    start_column = 12  # Column K

    # Iterate through the list and populate cells
    for value in values_list:
        cell = NewSheetEdit.cell(row=start_row, column=start_column)
        cell.value = value

        # Populate the two cells below the current cell
        for i in range(1, 3):
            cell_below = NewSheetEdit.cell(row=start_row + i, column=start_column)
            cell_below.value = value
        
        start_row += 3  # Move to the next group of cells

else:
    MRAdder = MRLow
    MRIndex = 0
    MRIntArray = []  # Initialize an empty list
    while MRAdder <= MRHigh:
        MRIntArray.append(MRAdder)  # Append MRAdder to the list
        MRAdder = round(MRAdder+MRInt,3)
        MRIndex += 1
    values_list = MRIntArray  # Replace with your list of values

    # Define the starting row and column
    start_row = 3  # Row K3
    start_column = 12  # Column K

    # Iterate through the list and populate cells
    for value in values_list:
        cell = NewSheetEdit.cell(row=start_row, column=start_column)
        cell.value = value
        # Populate the two cells below the current cell
        for i in range(1, 3):
            cell_below = NewSheetEdit.cell(row=start_row + i, column=start_column)
            cell_below.value = value
        start_row += 3  # Move to the next group of cells

# GRAPH TIME

# Isp over MR Graph

MR_x_data = values_list

# Handles ISP at exit
IspRowCount = 5
Isp_y_data = []
IspCellValue = 1
gravity = 9.80665
while IspCellValue != None:
    IspCell = NewSheetEdit[f"R{IspRowCount}"]
    IspCellValue = IspCell.value
    IspRowCount += 3
    if IspCellValue != None and len(Isp_y_data) != len(MR_x_data):
        Isp_y_data.append(IspCellValue/gravity) 

# Create the scatterplot using matplotlib
plt.scatter(MR_x_data, Isp_y_data, label='Isp over MR1')
plt.xlabel('Mixture Ratio')
plt.ylabel('Isp')
plt.title('Isp over MR2')
plt.legend()

# Customize the appearance of the scatterplot as needed
# For example:
# plt.grid(True)
# plt.xlim(0, 10)
# plt.ylim(0, 20)

# Save the scatterplot to a file (e.g., 'scatterplot.png')
#plt.savefig('CEAProcessing/scatterplot.png')

# Display the scatterplot
#plt.show()

# data

x_data = MR_x_data
y_data = Isp_y_data

# Write data to worksheet
for i, (x, y) in enumerate(zip(x_data, y_data), start=2):
    NewSheetEdit[f'V{i}'] = x
    NewSheetEdit[f'W{i}'] = y

# Create a scatter chart
chart = ScatterChart()
chart.title = "Isp over Mixture Ratio"
chart.style = 15  # Set the chart style (you can change this as needed)

# Define references for x and y values
x_values = Reference(NewSheetEdit, min_col=22, min_row=2, max_row=len(x_data) + 1)
y_values = Reference(NewSheetEdit, min_col=23, min_row=2, max_row=len(y_data) + 1)

# Create a series for the chart with smooth lines set to False
series = Series(y_values, x_values)
series.graphicalProperties.line.noFill=True
series.marker.symbol = "circle"
series.trendline = Trendline(dispEq=True, trendlineType = 'poly', order = 5)

# Add the series to the chart
chart.series.append(series)
chart.legend = None

# Access the x-axis and y-axis
x_axis = chart.x_axis
y_axis = chart.y_axis

# Configure major ticks on the x-axis
x_axis.majorTickMark = "in"  # Options: "none", "in", "out", "cross"
x_axis.majorUnit = .1  # Set the interval between major ticks

# Add axis labels
chart.x_axis.title = "Mixture Ratio"
chart.y_axis.title = "Isp"

# Set x-axis limit
if MRBox == True:
    chart.x_axis.scaling.min = (min(values_list)-0.1)  # Change this value as needed
    chart.x_axis.scaling.max = (max(values_list)+0.1)  # Change this value as needed
else:
    chart.x_axis.scaling.min = MRLow-0.1  # Change this value as needed
    chart.x_axis.scaling.max = MRHigh+0.1  # Change this value as needed

#Size
chart.height = 10.05 # default is 7.5
chart.width = 20.2 # default is 15

# Add the chart to the worksheet
NewSheetEdit.add_chart(chart, "V2")

#Temp over MR Chart

# Handles Temp at chamber
TempRowCount = 3
Temp_y_data = []
TempCellValue = 1
while TempCellValue != None:
    TempCell = NewSheetEdit[f"Q{TempRowCount}"]
    TempCellValue = TempCell.value
    TempRowCount += 3
    if TempCellValue != None:
        Temp_y_data.append(TempCellValue) 

# data
x_data = MR_x_data
y_data = Temp_y_data

# Write data to worksheet
for i, (x, y) in enumerate(zip(x_data, y_data), start=33):
    NewSheetEdit[f'V{i}'] = x
    NewSheetEdit[f'W{i}'] = y

# Create a scatter chart
chart = ScatterChart()
chart.title = "Chamber Temp over Mixture Ratio"
chart.style = 15  # Set the chart style (you can change this as needed)

# Define references for x and y values
x_values = Reference(NewSheetEdit, min_col=22, min_row=33, max_row=len(x_data) + 32)
y_values = Reference(NewSheetEdit, min_col=23, min_row=33, max_row=len(y_data) + 32)

# Create a series for the chart with smooth lines set to False
series = Series(y_values, x_values)
series.graphicalProperties.line.noFill=True
series.marker.symbol = "circle"
series.trendline = Trendline(dispEq=True, trendlineType = 'poly', order = 5)

# Add the series to the chart
chart.series.append(series)
chart.legend = None

# Access the x-axis and y-axis
x_axis = chart.x_axis
y_axis = chart.y_axis

# Configure major ticks on the x-axis
x_axis.majorTickMark = "in"  # Options: "none", "in", "out", "cross"
x_axis.majorUnit = .1  # Set the interval between major ticks

# Add axis labels
chart.x_axis.title = "Mixture Ratio"
chart.y_axis.title = "Chamber Temp (K)"

# x limits
if MRBox == True:
    chart.x_axis.scaling.min = (min(values_list)-0.1)  # Change this value as needed
    chart.x_axis.scaling.max = (max(values_list)+0.1)  # Change this value as needed
else:
    chart.x_axis.scaling.min = MRLow-0.1  # Change this value as needed
    chart.x_axis.scaling.max = MRHigh+0.1  # Change this value as needed

# Set y-axis limit
chart.y_axis.scaling.min = 100*(round(min(Temp_y_data)/100))-200  # Change this value as needed
chart.y_axis.scaling.max = 100*(round(max(Temp_y_data)/100))+200  # Change this value as needed

#Size
chart.height = 10.05 # default is 7.5
chart.width = 20.2 # default is 15

# Add the chart to the worksheet
NewSheetEdit.add_chart(chart, "V21")

#Change V and W columns text to white

# Define the columns you want to change the text color for (columns V and W)
columns_to_change = ["V", "W"]

# NO MO CHART TIME

# Set the font color to white for the specified columns
font = Font(color="FFFFFF")  # White color

for col in columns_to_change:
    for cell in NewSheetEdit[col]:
        cell.font = font

NewSheetEdit['X41'] = f"According to CEA Data (not trendline), the Max Isp is {round((max(Isp_y_data)),2)} at an MR of {values_list[Isp_y_data.index(max(Isp_y_data))]}."
NewSheetEdit['X42'] = f"This corresponds to a chamber temperature of {round(Temp_y_data[Isp_y_data.index(max(Isp_y_data))],2)} K."

# Define border styles
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
left_border = Border(left=Side(style='thin'))
right_border = Border(right=Side(style='thin'))
top_border = Border(top=Side(style='thin'))
bottom_border = Border(bottom=Side(style='thin'))

# Apply the specified border styles to the cells
# Apply right border to cells W41 and W42
NewSheetEdit['W41'].border = right_border
NewSheetEdit['W42'].border = right_border

# Apply left border to cells AF41 and AF42
NewSheetEdit['AF41'].border = left_border
NewSheetEdit['AF42'].border = left_border

# Apply bottom border to cells X40 to AE40
for cell in NewSheetEdit['X40:AE40'][0]:
    cell.border = bottom_border

# Apply top border to cells X43 to AE43
for cell in NewSheetEdit['X43:AE43'][0]:
    cell.border = top_border

# End Sheet Edit

# Kill Selenium
driver.close()

print("'AutoCEA.py' complete, closing browser and opening 'AutoCEA_Output.xlsx'. \U0001F680\n\n")

OutputExcelFile.save('CEAProcessing\AutoCEA_Output.xlsx')
OutputExcelFile.close()
os.system('CEAProcessing\AutoCEA_Output.xlsx')